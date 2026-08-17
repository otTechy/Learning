import pandas as pd
from xbbg import blp
from datetime import datetime
import argparse
import numpy as np
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

class Risk:
    def __init__(self, portfolio_returns, benchmark_returns=None, risk_free_rate=0.02):
        """
        :param portfolio_returns: pd.Series of monthly portfolio returns
        :param benchmark_returns: pd.Series of monthly benchmark returns (optional)
        :param risk_free_rate: annual risk-free rate (e.g., 0.02 for 2%)
        """
        self.returns = portfolio_returns
        self.benchmark_returns = benchmark_returns
        self.risk_free_rate = risk_free_rate / 12  # Convert to monthly
        self.metrics = {}

    def calculate_volatility(self):
        return self.returns.std()

    def calculate_max_drawdown(self):
        cumulative = (1 + self.returns).cumprod()
        peak = cumulative.cummax()
        drawdown = (cumulative - peak) / peak
        return drawdown.min()

    def calculate_beta(self):
        if self.benchmark_returns is not None:
            cov_matrix = np.cov(self.returns, self.benchmark_returns)
            cov = cov_matrix[0][1]
            var = np.var(self.benchmark_returns)
            return cov / var if var != 0 else np.nan
        return None

    def calculate_var(self, confidence_level=0.95):
        return np.percentile(self.returns, (1 - confidence_level) * 100)

    def calculate_sharpe_ratio(self):
        excess_return = self.returns - self.risk_free_rate
        return excess_return.mean() / self.returns.std()

    def calculate_sortino_ratio(self):
        excess_return = self.returns - self.risk_free_rate
        downside_std = excess_return[excess_return < 0].std()
        return excess_return.mean() / downside_std if downside_std != 0 else np.nan

    def evaluate(self):
        self.metrics['Volatility'] = self.calculate_volatility()
        self.metrics['Max Drawdown'] = self.calculate_max_drawdown()
        self.metrics['Beta'] = self.calculate_beta()
        self.metrics['VaR (95%)'] = self.calculate_var()
        self.metrics['Sharpe Ratio'] = self.calculate_sharpe_ratio()
        self.metrics['Sortino Ratio'] = self.calculate_sortino_ratio()
        return pd.Series(self.metrics)

class ETFPortfolio:
    def __init__(self, tickers, weights, benchmarks, start_date, end_date):
        self.tickers = tickers
        self.weights = weights
        self.benchmarks = benchmarks
        self.start_date = start_date
        self.end_date = end_date
        self.bbg_tickers = [f"{tkr} US Equity" for tkr in self.tickers]
        self.bench_tickers = list(self.benchmarks.values())
        self.all_tickers = self.bbg_tickers + self.bench_tickers
        self.adj_close = None
        self.performance_df = None

    def fetch_data(self):
        data = blp.bdh(self.all_tickers, flds=['PX_LAST'], start_date=self.start_date, end_date=self.end_date, Per='M')
        self.adj_close = data.xs('PX_LAST', axis=1, level=1)
        self.adj_close.columns = self.tickers + list(self.benchmarks.keys())
        self.adj_close.dropna(inplace=True)

    def calculate_performance(self):
        monthly_returns = self.adj_close[self.tickers].pct_change().dropna()
        weights_series = pd.Series(self.weights, index=self.tickers)
        portfolio_returns = monthly_returns.dot(weights_series)
        cumulative_returns = (1 + portfolio_returns).cumprod()
        bench_returns = self.adj_close[list(self.benchmarks.keys())].pct_change().dropna()
        bench_cum_returns = (1 + bench_returns).cumprod()
        perf_dict = {
            "Portfolio Monthly Return": portfolio_returns,
            "Portfolio Cumulative Return": cumulative_returns
        }
        for b in self.benchmarks.keys():
            perf_dict[f"{b} Monthly Return"] = bench_returns[b]
            perf_dict[f"{b} Cumulative Return"] = bench_cum_returns[b]
        self.performance_df = pd.DataFrame(perf_dict)
        return self.performance_df

    def save_to_excel(self, filename='etf_portfolio_performance.xlsx', risk_df=None):
        if self.performance_df is not None:
            self.performance_df.to_excel(filename, index=True)
            print(f'Results saved to {filename}')
            wb = openpyxl.load_workbook(filename)
            # Save risk metrics to a new sheet
            if risk_df is not None:
                if 'Risk Metrics' in wb.sheetnames:
                    ws = wb['Risk Metrics']
                    wb.remove(ws)
                ws = wb.create_sheet('Risk Metrics')
                for r in dataframe_to_rows(risk_df, index=True, header=True):
                    ws.append(r)
                print('Risk metrics saved to Risk Metrics tab in Excel.')
            # Save chart to a new sheet
            if 'Chart' in wb.sheetnames:
                ws = wb['Chart']
                wb.remove(ws)
            chart_sheet = wb.create_sheet('Chart')
            chart_data = self.performance_df[[col for col in self.performance_df.columns if 'Monthly Return' in col]]
            for r in dataframe_to_rows(chart_data, index=True, header=True):
                chart_sheet.append(r)
            chart = LineChart()
            chart.title = "Portfolio & Benchmark Monthly Returns"
            chart.y_axis.title = "Monthly Return"
            chart.x_axis.title = "Date"
            # Show all dates as x-axis labels
            chart.x_axis.majorTickMark = "in"
            chart.x_axis.tickLblSkip = 1
            chart.x_axis.label_rotation = 45
            # Set y-axis reasonable interval (auto min/max, step 0.05)
            chart.y_axis.majorUnit = 0.05
            # Add data and categories
            data = Reference(chart_sheet, min_col=2, max_col=1+len(chart_data.columns), min_row=1, max_row=1+len(chart_data))
            cats = Reference(chart_sheet, min_col=1, min_row=2, max_row=1+len(chart_data))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 12
            chart.width = 24
            # Legend at the bottom
            chart.legend.position = "b"
            chart_sheet.add_chart(chart, "A20")
            wb.save(filename)
            print('Chart added to Chart tab in Excel.')
        else:
            print('No performance data to save.')

    def calculate_risk_metrics(self):
        """
        Calculate risk metrics for the portfolio and each benchmark using the Risk class.
        Returns a DataFrame with risk metrics for portfolio and benchmarks.
        """
        risk_results = {}
        # Portfolio risk
        portfolio_returns = self.performance_df["Portfolio Monthly Return"]
        risk_results['Portfolio'] = Risk(portfolio_returns).evaluate()
        # Benchmark risks
        for b in self.benchmarks.keys():
            bench_returns = self.performance_df.get(f"{b} Monthly Return")
            if bench_returns is not None:
                risk_results[b] = Risk(portfolio_returns, bench_returns).evaluate()
        return pd.DataFrame(risk_results)

def parse_list(str_value):
    return [item.strip() for item in str_value.split(',') if item.strip()]

def parse_weights(str_value):
    return [float(item) for item in str_value.split(',') if item.strip()]

def parse_benchmarks(str_value):
    # Accepts comma-separated tickers, e.g. "SPMARC5P Index,SPXFRRE7 Index"
    return {b: b for b in [item.strip() for item in str_value.split(',') if item.strip()]}

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--tickers', type=str, required=True, help='Comma-separated ETF tickers (e.g. SPY,VXUS,QQQ)')
    parser.add_argument('--weights', type=str, required=True, help='Comma-separated weights (e.g. 0.2,0.15,0.1)')
    parser.add_argument('--benchmarks', type=str, required=True, help='Comma-separated benchmarks as name=ticker (e.g. SPMARC5P Index=SPMARC5P Index,SPXFRRE7 Index=SPXFRRE7 Index)')
    parser.add_argument('--start_date', type=str, required=True)
    parser.add_argument('--end_date', type=str, required=True)
    args = parser.parse_args()

    tickers = parse_list(args.tickers)
    weights = parse_weights(args.weights)
    if len(tickers) != len(weights):
        raise ValueError('Number of tickers and weights must match!')
    benchmarks = parse_benchmarks(args.benchmarks)

    portfolio = ETFPortfolio(tickers, weights, benchmarks, args.start_date, args.end_date)
    portfolio.fetch_data()
    performance_df = portfolio.calculate_performance()
    print(performance_df)
    # Calculate and print risk metrics for portfolio and all benchmarks
    risk_df = portfolio.calculate_risk_metrics()
    print("\nRisk Metrics (Portfolio and Benchmarks):")
    print(risk_df)
    portfolio.save_to_excel(risk_df=risk_df)
    return performance_df

if __name__ == "__main__":
    main()

